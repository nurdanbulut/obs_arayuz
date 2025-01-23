from openpyxl import Workbook, load_workbook

def get_user_inputs(file_name, prompt):
    """
    Kullanıcıdan sayıyı ve string değerleri alıp bir dosyaya yazan işlev.
    """
    while True:
        try:
            count = int(input(f"Kaç {prompt} gireceksiniz? "))
            if count <= 0:
                print("Lütfen pozitif bir sayı girin.")
                continue
            break
        except ValueError:
            print("Geçerli bir sayı girin.")

    inputs = []
    for i in range(count):
        value = input(f"{i+1}. {prompt}: ")
        inputs.append(value)

    with open(file_name, "w", encoding="utf-8") as f:
        f.write("\n".join(inputs))

    print(f"{file_name} dosyasına başarıyla kaydedildi.")

def create_table1(prg_cikti_dosya, ders_cikti_dosya, file_name):
    # Program ve ders çıktıları dosyalarını oku
    with open(prg_cikti_dosya, "r", encoding="utf-8") as f:
        prg_cikti = [line.strip() for line in f if line.strip()]

    with open(ders_cikti_dosya, "r", encoding="utf-8") as f:
        ders_cikti = [line.strip() for line in f if line.strip()]
    global ders_ciktisi_sayisi
    program_ciktisi_sayisi = len(prg_cikti)
    ders_ciktisi_sayisi = len(ders_cikti)

    # Excel çalışma kitabını oluştur
    wb = Workbook()
    ws1 = wb.create_sheet("Tablo 1", 0)

    # Başlıkları yaz
    ws1.cell(row=1, column=1, value="Program Çıktıları")
    for col in range(2, ders_ciktisi_sayisi + 2):
        ws1.cell(row=1, column=col, value=f"Ders Çıktı{col - 1}")
    ws1.cell(row=1, column=ders_ciktisi_sayisi + 2, value="İlişki Değerlendirme")

    # Program çıktıları ve kullanıcıdan alınan verilerle tabloyu doldur
    for row in range(2, program_ciktisi_sayisi + 2):
        # Program çıktısı ismini yaz
        ws1.cell(row=row, column=1, value=prg_cikti[row - 2])

        ders_cikti_toplam = 0  # Ağırlık hesaplama için toplam

        for col in range(2, ders_ciktisi_sayisi + 2):
            while True:
                try:
                    value = float(input(f"{prg_cikti[row - 2]} için {ders_cikti[col - 2]} ilişkisini girin (0 ile 1 arasında): "))
                    if 0 <= value <= 1:
                        ws1.cell(row=row, column=col, value=value)
                        ders_cikti_toplam += value
                        break
                    else:
                        print("Hata: Değer 0 ile 1 arasında olmalı!")
                except ValueError:
                    print("Hata: Geçerli bir sayı girin!")

        # İlişki Değerlendirme sütununu hesapla
        if ders_ciktisi_sayisi > 0:
            ws1.cell(row=row, column=ders_ciktisi_sayisi + 2, value=round(ders_cikti_toplam / ders_ciktisi_sayisi, 2))
        else:
            ws1.cell(row=row, column=ders_ciktisi_sayisi + 2, value=0)

    # Çalışma kitabını kaydet
    wb.save(file_name)
    print(f"Tablo 1 {file_name} dosyasına kaydedildi.")

def create_table2(file_name):
    # Dosyayı yükle
    wb = load_workbook(file_name)

    # Tablo 1'i oku ve Ders Çıktısı sayısını belirle
    ws1 = wb["Tablo 1"]
    ders_ciktisi_sayisi = ws1.max_column - 2  # İlk sütun "Program Çıktıları", son sütun "İlişki Değerlendirme"

    # Tablo 2'yi oluştur
    if "Tablo 2" in wb.sheetnames:
        ws2 = wb["Tablo 2"]
    else:
        ws2 = wb.create_sheet("Tablo 2")

    # Başlıkları oluştur
    basliklar = ["Ders Çıktı", "Ödev1", "Ödev2", "Quiz", "Vize", "Final", "Toplam"]
    yuzdelikler = {}

    def toplam_yuzde_kontrol(yuzdelikler):
        return sum(yuzdelikler.values()) == 100

    # Kullanıcıdan yüzde bilgilerini al
    while True:
        yuzdelikler.clear()
        toplam_yuzde = 0

        for col, criterion in enumerate(basliklar[1:-1], start=2):  # "Ders Çıktı" ve "Toplam" dışındaki başlıklar
            while True:
                try:
                    yuzde = float(input(f"{criterion} için yüzde etkisini girin (%): "))
                    if 0 <= yuzde <= 100:
                        yuzdelikler[criterion] = yuzde
                        toplam_yuzde += yuzde
                        ws2.cell(row=1, column=col, value=f"{criterion} (%{yuzde})")  # Yüzdelik bilgisi başlığa eklenir
                        break
                    else:
                        print("Hata: Yüzde değeri 0 ile 100 arasında olmalıdır.")
                except ValueError:
                    print("Hata: Geçerli bir sayı girin!")

        if toplam_yuzde_kontrol(yuzdelikler):
            break
        else:
            print("Hata: Yüzdelerin toplamı 100 olmalıdır. Lütfen tekrar girin!")

    # Kullanıcının girdiği yüzdelikleri kontrol etmek için çıktı veriyoruz.
    print("Girilen yüzdelikler:", yuzdelikler)
    print("Toplam yüzde:", sum(yuzdelikler.values()))

    # "Ders Çıktı" başlığını yaz
    ws2.cell(row=1, column=1, value="Ders Çıktı")

    # Kullanıcıdan Ders Çıktıları için veri al
    for row in range(2, ders_ciktisi_sayisi + 2):
        ws2.cell(row=row, column=1, value=f"DC{row - 1}")  # Ders Çıktı ismi
        toplam_iliski = 0  # Toplam ilişkiyi hesaplamak için

        for col, criterion in enumerate(basliklar[1:-1], start=2):  # "Ders Çıktı" ve "Toplam" dışındaki başlıklar
            while True:
                try:
                    iliski = int(input(f"{ws2.cell(row=row, column=1).value} için {criterion} ile ilişkisi (0 veya 1): "))
                    if iliski in [0, 1]:
                        ws2.cell(row=row, column=col, value=iliski)
                        toplam_iliski += iliski  # Toplam ilişkiyi artır
                        break
                    else:
                        print("Hata: Değer sadece 0 veya 1 olmalı!")
                except ValueError:
                    print("Hata: Geçerli bir sayı girin!")

        # Toplam sütununa yaz
        ws2.cell(row=row, column=7, value=toplam_iliski)

    # Çalışma kitabını kaydet
    wb.save(file_name)
    print(f"Tablo 2 {file_name} dosyasına kaydedildi.")


# Kullanım
prg_cikti_dosya = "prg_cikti_dosyasi.txt"  # Program çıktıları dosyası
ders_cikti_dosya = "ders_cikti_dosyasi.txt"  # Ders çıktıları dosyası
file_name = "degerlendirme.xlsx"  # Excel dosyası

# Kullanıcıdan çıktıları al ve dosyalara kaydet
get_user_inputs(prg_cikti_dosya, "program çıktısı")
get_user_inputs(ders_cikti_dosya, "ders çıktısı")

# Tablo 1 ve Tablo 2'yi oluştur
create_table1(prg_cikti_dosya, ders_cikti_dosya, file_name)
create_table2(file_name)


def create_table3(file_name):
    # Dosyayı yükle
    wb = load_workbook(file_name)

    # Tablo 2 ve değerlendirme kriterlerini oku
    ws2 = wb["Tablo 2"]
    ders_ciktisi_sayisi = ws2.max_row - 1  # "Ders Çıktı" başlığını hariç tutuyoruz
    kriter_sayisi = ws2.max_column - 2  # "Ders Çıktı" ve "Toplam" hariç sütun sayısı

    # Tablo 3'ü oluştur
    if "Tablo 3" in wb.sheetnames:
        ws3 = wb["Tablo 3"]
    else:
        ws3 = wb.create_sheet("Tablo 3")

    # Tablo 3 başlıklarını yaz
    ws3.cell(row=1, column=1, value="Tablo 3")
    for col in range(2, kriter_sayisi + 3):  # Değerlendirme kriterleri ve Toplam sütunu
        ws3.cell(row=1, column=col, value=ws2.cell(row=1, column=col).value)

    # Değerlendirme kriterleri yüzdelerini al (Tablo 2'nin başlıklarındaki yüzdelikler)
    yuzdelikler = {}
    for col in range(2, kriter_sayisi + 2):
        baslik = ws2.cell(row=1, column=col).value
        try:
            yuzdelik = float(baslik.split("(%")[1].replace("%)", "").replace(")", "")) / 100  # "%10.0)" -> 0.1
        except (IndexError, ValueError):
            raise ValueError(f"Başlık formatı hatalı: '{baslik}'. Başlık, '(%)' formatında olmalı.")
        yuzdelikler[col] = yuzdelik

    # Tablo 3'ü doldur
    for row in range(2, ders_ciktisi_sayisi + 2):
        # Ders çıktısı adını yaz
        ws3.cell(row=row, column=1, value=ws2.cell(row=row, column=1).value)

        toplam_agirlikli_deger = 0
        for col in range(2, kriter_sayisi + 2):
            iliski = ws2.cell(row=row, column=col).value or 0
            agirlik = yuzdelikler[col] * iliski
            ws3.cell(row=row, column=col, value=round(agirlik, 2))
            toplam_agirlikli_deger += agirlik

        # Toplam sütununa yaz
        ws3.cell(row=row, column=kriter_sayisi + 2, value=round(toplam_agirlikli_deger, 2))

    # Çalışma kitabını kaydet
    wb.save(file_name)
    print(f"Tablo 3 {file_name} dosyasına kaydedildi.")


# Kullanım
file_name = "degerlendirme.xlsx"
create_table3(file_name)


def create_table4(grades_file, weights_file):
    # Dosyaları yükle
    grades_wb = load_workbook(grades_file, data_only=True)
    weights_wb = load_workbook(weights_file)

    # Çalışma sayfalarını seç
    grades_ws = grades_wb.active
    weights_ws = weights_wb['Tablo 3']

    # Tablo 4'ü oluştur
    if 'Tablo 4' in weights_wb.sheetnames:
        del weights_wb['Tablo 4']
    table4_ws = weights_wb.create_sheet('Tablo 4')

    # Tablo 3'ten ağırlıklar ve max değerleri oku
    weights = {}
    max_values = {}
    for row in range(2, weights_ws.max_row + 1):
        ders_cikti = weights_ws.cell(row=row, column=1).value
        row_weights = {
            'Odev1': weights_ws.cell(row=row, column=2).value or 0,
            'Odev2': weights_ws.cell(row=row, column=3).value or 0,
            'Quiz': weights_ws.cell(row=row, column=4).value or 0,
            'Vize': weights_ws.cell(row=row, column=5).value or 0,
            'Final': weights_ws.cell(row=row, column=6).value or 0
        }
        weights[ders_cikti] = row_weights
        max_values[ders_cikti] = (weights_ws.cell(row=row, column=7).value or 0) * 100

    # Başlıklar
    headers = ['Ders Çıktı', 'Odev1', 'Odev2', 'Quiz', 'Vize', 'Final', 'TOPLAM', 'MAX', '% Başarı']

    # Her öğrenci için Tablo 4'ü doldur
    current_row = 1
    for student_row in range(2, grades_ws.max_row + 1):
        student_no = grades_ws.cell(row=student_row, column=1).value

        # Öğrenci başlığı yaz
        table4_ws.cell(row=current_row, column=1, value="TABLO 4")
        table4_ws.cell(row=current_row, column=2, value=f"Öğrenci {student_no} için")
        current_row += 1

        # Sütun başlıklarını yaz
        for col, header in enumerate(headers, 1):
            table4_ws.cell(row=current_row, column=col, value=header)
        current_row += 1

        # Öğrencinin notlarını al
        grades = {
            'Odev1': grades_ws.cell(row=student_row, column=2).value or 0,
            'Odev2': grades_ws.cell(row=student_row, column=3).value or 0,
            'Quiz': grades_ws.cell(row=student_row, column=4).value or 0,
            'Vize': grades_ws.cell(row=student_row, column=5).value or 0,
            'Final': grades_ws.cell(row=student_row, column=6).value or 0
        }

        # Her ders çıktısı için hesaplama yap
        for ders_cikti in weights.keys():
            weighted_total = 0
            for category, grade in grades.items():
                weighted_total += grade * weights[ders_cikti][category]

            # Tablo 4'e yaz
            table4_ws.cell(row=current_row, column=1, value=ders_cikti)
            table4_ws.cell(row=current_row, column=2, value=round(grades['Odev1'] * weights[ders_cikti]['Odev1'], 2))
            table4_ws.cell(row=current_row, column=3, value=round(grades['Odev2'] * weights[ders_cikti]['Odev2'], 2))
            table4_ws.cell(row=current_row, column=4, value=round(grades['Quiz'] * weights[ders_cikti]['Quiz'], 2))
            table4_ws.cell(row=current_row, column=5, value=round(grades['Vize'] * weights[ders_cikti]['Vize'], 2))
            table4_ws.cell(row=current_row, column=6, value=round(grades['Final'] * weights[ders_cikti]['Final'], 2))
            table4_ws.cell(row=current_row, column=7, value=round(weighted_total, 2))
            table4_ws.cell(row=current_row, column=8, value=round(max_values[ders_cikti], 2))

            # Başarı yüzdesini hesapla
            success_rate = (weighted_total / max_values[ders_cikti] * 100) if max_values[ders_cikti] > 0 else 0
            table4_ws.cell(row=current_row, column=9, value=round(success_rate, 1))

            current_row += 1

        # Bir sonraki öğrenci için boşluk bırak
        current_row += 2

    # Dosyayı kaydet
    weights_wb.save(weights_file)
    print("Tablo 4 oluşturuldu ve kaydedildi.")


# Kullanım
grades_file = "NotYukle-BLM315-2024-1 (1).xlsx"
weights_file = "degerlendirme.xlsx"
create_table4(grades_file, weights_file)


def create_table5(weights_file):
    # Dosyayı yükle
    wb = load_workbook(weights_file)

    # Tablo 1 ve Tablo 4'ü yükle
    table1_ws = wb["Tablo 1"]
    table4_ws = wb["Tablo 4"]

    # Tablo 5'i oluştur
    if "Tablo 5" in wb.sheetnames:
        del wb["Tablo 5"]
    table5_ws = wb.create_sheet("Tablo 5")

    current_row = 1
    # Tablo 4'teki her öğrenci grubu için
    row = 1
    while row <= table4_ws.max_row:
        # Öğrenci başlığını bul
        student_header = table4_ws.cell(row=row, column=2).value
        if student_header and "Öğrenci" in student_header:
            # Tablo 5 başlığını yaz
            table5_ws.cell(row=current_row, column=1, value="TABLO 5")
            table5_ws.cell(row=current_row, column=2, value=student_header)
            current_row += 1

            # Ders çıktısı başlığını yaz
            table5_ws.cell(row=current_row, column=2, value="Ders çıktısı")
            table5_ws.cell(row=current_row, column=8, value="Başarı Oranı")
            current_row += 1

            # Başarı oranlarını DC ile eşleştir
            success_rates = {}
            row += 2  # Başlıkları atla

            while row <= table4_ws.max_row and not (table4_ws.cell(row=row, column=2).value and "Öğrenci" in str(table4_ws.cell(row=row, column=2).value)):
                dc = table4_ws.cell(row=row, column=1).value
                if dc and dc.startswith("DC"):
                    success_rate = table4_ws.cell(row=row, column=9).value
                    success_rates[dc] = success_rate
                row += 1

            # Program çıktılarını yaz
            for prog_row in range(2, table1_ws.max_row + 1):
                prog_output = table1_ws.cell(row=prog_row, column=1).value
                table5_ws.cell(row=current_row, column=1, value=prog_output)

                # İlgili DC başarı oranlarını bul ve toplama ekle
                dc_success_total = 0
                dc_count = 0
                for dc_col in range(2, table1_ws.max_column):
                    relation = table1_ws.cell(row=prog_row, column=dc_col).value
                    if relation == 1:  # İlgili DC
                        dc_number = f"DC{dc_col - 1}"
                        success = success_rates.get(dc_number, 0)
                        table5_ws.cell(row=current_row, column=dc_col, value=success)
                        dc_success_total += success
                        dc_count += 1
                    else:
                        table5_ws.cell(row=current_row, column=dc_col, value=0)

                # Tablo 1'deki "İlişki Değerlendirme" sütununu al
                relation_evaluation = table1_ws.cell(row=prog_row, column=table1_ws.max_column).value or 1

                # Başarı oranını hesapla
                if relation_evaluation > 0:
                    avg_success = (dc_success_total/ders_ciktisi_sayisi) / relation_evaluation
                else:
                    avg_success = 0
                table5_ws.cell(row=current_row, column=8, value=round(avg_success, 1))

                current_row += 1
            current_row += 2  # Boşluk bırak
        else:
            row += 1

    # Dosyayı kaydet
    wb.save(weights_file)
    print("Tablo 5 oluşturuldu ve kaydedildi.")

create_table5("degerlendirme.xlsx")