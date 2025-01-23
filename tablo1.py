from PyQt5.QtCore import Qt
from openpyxl import Workbook
from PyQt5.QtWidgets import (
    QApplication, QWidget, QVBoxLayout, QLabel, QPushButton, QMessageBox, QTableWidget, QTableWidgetItem, QHeaderView
)
import sys

class Tablo1App(QWidget):
    def __init__(self):
        super().__init__()
        self.initUI()

    def initUI(self):
        # Pencerenin başlığını ve boyutunu ayarla
        self.setWindowTitle("Tablo 1 Oluşturma ve Hesaplama")
        self.setGeometry(100, 100, 800, 600)

        # Ana dikey yerleşim düzenini oluştur
        self.layout = QVBoxLayout()

        # Tablo oluşturuluyor (5 satır, 5 sütun)
        self.table = QTableWidget(5, 5)
        # Sütun başlıkları ayarlanıyor
        self.table.setHorizontalHeaderLabels(["Program Çıktıları"] + [f"Ders Çıktı {i+1}" for i in range(3)] + ["İlişki Değeri"])
        # Sütun genişliklerini otomatik olarak yerleştir
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        # Tablo hücrelerinde değişiklik olduğunda kontrol et
        self.table.itemChanged.connect(self.validate_and_update)
        self.layout.addWidget(self.table)

        # Program çıktıları sütununun ilk değerlerini belirle
        for row in range(self.table.rowCount()):
            self.table.setItem(row, 0, QTableWidgetItem(str(row + 1)))
            # İlk sütunu düzenlenemez hale getir
            self.table.item(row, 0).setFlags(Qt.ItemIsEnabled)

        # Satır ekleme butonu
        self.add_row_button = QPushButton("Satır Ekle")
        self.add_row_button.clicked.connect(self.addRow)
        self.layout.addWidget(self.add_row_button)

        # Sütun ekleme butonu
        self.add_column_button = QPushButton("Sütun Ekle")
        self.add_column_button.clicked.connect(self.addColumn)
        self.layout.addWidget(self.add_column_button)

        # Tabloyu kaydetme butonu
        self.save_button = QPushButton("Tabloyu Kaydet")
        self.save_button.clicked.connect(self.saveTable)
        self.layout.addWidget(self.save_button)

        # Ana yerleşimi pencereye ekle
        self.setLayout(self.layout)

    def addRow(self):
        # Yeni bir satır eklemek için fonksiyon
        current_row_count = self.table.rowCount()
        self.table.insertRow(current_row_count)
        # Yeni satıra sıralı numara ekleniyor
        self.table.setItem(current_row_count, 0, QTableWidgetItem(str(current_row_count + 1)))
        self.table.item(current_row_count, 0).setFlags(Qt.ItemIsEnabled)

    def addColumn(self):
        # Yeni bir sütun eklemek için fonksiyon
        current_col_count = self.table.columnCount()
        self.table.insertColumn(current_col_count - 1)
        # Yeni sütuna başlık ekleniyor
        self.table.setHorizontalHeaderItem(current_col_count - 1, QTableWidgetItem(f"Ders Çıktı {current_col_count - 1}"))

    def validate_and_update(self, item):
        # Kullanıcının girdisini kontrol et
        if item.column() == 0 or item.column() == self.table.columnCount() - 1:
            return

        try:
            # Girdi 0 ile 1 arasında olmalıdır
            value = float(item.text())
            if value < 0 or value > 1:
                raise ValueError
        except ValueError:
            # Hatalı giriş durumunda uyarı ver
            QMessageBox.warning(self, "Hata", "Lütfen 0 ile 1 arasında bir değer girin.")
            item.setText("")  # Hatalı değer girildiğinde temizleniyor
            return

        self.updateRelations()

    def updateRelations(self):
        # İlişki değerlerini hesapla ve güncelle
        try:
            for row in range(self.table.rowCount()):
                total_relation = 0
                # İlk ve son sütun hariç diğer değerlerin toplamını hesapla
                for col in range(1, self.table.columnCount() - 1):
                    cell_item = self.table.item(row, col)
                    if cell_item and cell_item.text():
                        total_relation += float(cell_item.text())
                # Ortalama ilişki değerini hesapla
                avg_relation = total_relation / (self.table.columnCount() - 2)
                self.table.setItem(row, self.table.columnCount() - 1, QTableWidgetItem(f"{avg_relation:.2f}"))
        except Exception as e:
            QMessageBox.warning(self, "Hata", f"Hesaplama sırasında bir hata oluştu: {e}")

    def saveTable(self):
        # Tabloyu Excel dosyasına kaydetme
        file_name = "Tablo1.xlsx"

        try:
            row_count = self.table.rowCount()
            col_count = self.table.columnCount()

            wb = Workbook()
            ws1 = wb.active
            ws1.title = "Tablo 1"

            # Başlık satırlarını Excel dosyasına yaz
            for col in range(col_count):
                header_item = self.table.horizontalHeaderItem(col)
                ws1.cell(row=1, column=col+1, value=header_item.text() if header_item else "")

            # Tablodaki tüm verileri Excel'e yaz
            for row in range(row_count):
                for col in range(col_count):
                    item = self.table.item(row, col)
                    value = item.text() if item else ""
                    ws1.cell(row=row+2, column=col+1, value=value)

            wb.save(file_name)
            QMessageBox.information(self, "Başarı", f"Tablo 1 {file_name} dosyasına kaydedildi.")
        except Exception as e:
            QMessageBox.critical(self, "Hata", f"Tablo kaydedilirken bir hata oluştu: {e}")


