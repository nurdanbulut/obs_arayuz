import sys
from PyQt5.QtWidgets import (
    QApplication, QWidget, QVBoxLayout, QLabel, QLineEdit, QPushButton, QFileDialog, QComboBox,
    QTableWidget, QTabWidget, QSpinBox, QFormLayout, QMessageBox
)
from PyQt5.QtCore import Qt
from tablo1 import Tablo1App
from tablo2 import Tablo2App
from openpyxl import load_workbook
#Not:Hocam proje dosyalarında "Tablo1.xlsx" dosyası varsa bu dosyayı silip kodu öyle çalıştırabilirsiniz.
class Proje3GUI(QWidget):
    def __init__(self):
        super().__init__()
        self.initUI()

    def initUI(self):
        # Pencere başlığını ve boyutunu ayarla
        self.setWindowTitle('Proje 3 Değerlendirme Arayüzü')

        self.main_layout = QVBoxLayout()
        self.resize(400, 500)

        # Ders seçimi için combo box
        self.ders_label = QLabel('Ders Seçin:')
        self.ders_combo = QComboBox()
        self.ders_combo.addItems(['YZM 315 - Yazılım Lab I', 'YZM 316 - Yazılım Lab II', 'YZM 317 - Yazılım Lab III'])

        # Program çıktısı dosyası seçimi
        self.prg_cikti_label = QLabel('Program Çıktıları İçin Dosya Seçimi(opsiyonel):')
        self.prg_cikti_input = QLineEdit()
        self.prg_cikti_button = QPushButton('Dosya Seç')
        self.style_button(self.prg_cikti_button)
        self.prg_cikti_button.clicked.connect(lambda: self.openFileDialog(self.prg_cikti_input))
        

        self.ders_cikti_label = QLabel('Ders Öğrenme Çıktıları:')
        self.ders_cikti_input = QLineEdit()
        self.ders_cikti_button = QPushButton('Dosya Seç')
        self.style_button(self.ders_cikti_button)
        self.ders_cikti_button.clicked.connect(lambda: self.openFileDialog(self.ders_cikti_input))

        self.add_tab_buttons()


        self.main_layout.addWidget(self.ders_label)
        self.main_layout.addWidget(self.ders_combo)
        self.main_layout.addWidget(self.prg_cikti_label)
        self.main_layout.addWidget(self.prg_cikti_input)
        self.main_layout.addWidget(self.prg_cikti_button)
        self.main_layout.addWidget(self.ders_cikti_label)
        self.main_layout.addWidget(self.ders_cikti_input)
        self.main_layout.addWidget(self.ders_cikti_button)


        self.setLayout(self.main_layout)

    #Butonlara özel stil
    def style_button(self, button):
        button.setStyleSheet("background-color: rgb(255, 192, 203); color: black; border-radius: 5px; padding: 5px;")
    #Dosya seçme diyalog penceresini açar.
    def openFileDialog(self, line_edit):
        options = QFileDialog.Options()
        file_name, _ = QFileDialog.getOpenFileName(self, 'Dosya Seç', '', 'Excel Dosyaları (*.xlsx);;Tüm Dosyalar (*)',
                                                   options=options)
        if file_name:
            line_edit.setText(file_name)

    #Tabloları açmak veya oluşturmak için butonları ekle
    def add_tab_buttons(self):
        self.tablo1_button = QPushButton('Tablo 1 Aç')
        self.style_button(self.tablo1_button)
        self.tablo1_button.clicked.connect(self.openTablo1)
        self.main_layout.addWidget(self.tablo1_button)

        self.tablo2_button = QPushButton('Tablo 2 Aç')
        self.style_button(self.tablo2_button)
        self.tablo2_button.clicked.connect(self.openTablo2)
        self.main_layout.addWidget(self.tablo2_button)

        self.tablo3_button = QPushButton('Tablo 3 Oluştur')
        self.style_button(self.tablo3_button)
        self.tablo3_button.clicked.connect(lambda: self.createTable3("Tablo1.xlsx"))
        self.main_layout.addWidget(self.tablo3_button)

        self.tablo4_button = QPushButton('Tablo 4 Oluştur')
        self.style_button(self.tablo4_button)
        self.tablo4_button.clicked.connect(lambda: self.createTable4("Tablo1.xlsx"))
        self.main_layout.addWidget(self.tablo4_button)

        self.tablo5_button = QPushButton('Tablo 5 Oluştur')
        self.style_button(self.tablo5_button)
        self.tablo5_button.clicked.connect(lambda: self.createTable5("Tablo1.xlsx"))
        self.main_layout.addWidget(self.tablo5_button)

    def openTablo1(self):
        self.tablo1_window = Tablo1App()
        self.tablo1_window.show()

    def openTablo2(self):
        self.tablo2_window = Tablo2App()
        self.tablo2_window.show()

    def createTable3(self, file_name):
        try:
            self._generate_table3(file_name)
            QMessageBox.information(self, "Başarı", "Tablo 3 başarıyla oluşturuldu ve kaydedildi.")
        except Exception as e:
            QMessageBox.critical(self, "Hata", f"Tablo 3 oluşturulurken bir hata oluştu: {e}")

    def _generate_table3(self, file_name):
        wb = load_workbook(file_name)
        ws2 = wb["Tablo 2"]
        ders_ciktisi_sayisi = ws2.max_row - 1
        kriter_sayisi = ws2.max_column - 2

        if "Tablo 3" in wb.sheetnames:
            del wb["Tablo 3"]
        ws3 = wb.create_sheet("Tablo 3")

        ws3.cell(row=1, column=1, value="Tablo 3")
        for col in range(2, kriter_sayisi + 3):
            ws3.cell(row=1, column=col, value=ws2.cell(row=1, column=col).value)

        yuzdelikler = {}
        for col in range(2, kriter_sayisi + 2):
            baslik = ws2.cell(row=1, column=col).value
            try:
                yuzde_parca = baslik.split("(%")[1].replace(")", "").strip()
                yuzdelik = float(yuzde_parca) / 100
            except (IndexError, ValueError):
                raise ValueError(f"Başlık formatı hatalı: '{baslik}'. Başlık '(%)' formatında olmalıdır.")
            yuzdelikler[col] = yuzdelik

        for row in range(2, ders_ciktisi_sayisi + 2):
            ws3.cell(row=row, column=1, value=ws2.cell(row=row, column=1).value)

            toplam_agirlikli_deger = 0
            for col in range(2, kriter_sayisi + 2):
                iliski = ws2.cell(row=row, column=col).value or 0
                if isinstance(iliski, (int, float)):
                    agirlik = yuzdelikler[col] * iliski
                else:
                    try:
                        agirlik = yuzdelikler[col] * float(iliski)
                    except ValueError:
                        raise ValueError(
                            f"Hatalı veri: Satır {row}, Sütun {col}. Beklenen bir sayı, ancak '{iliski}' bulundu.")
                ws3.cell(row=row, column=col, value=round(agirlik, 2))
                toplam_agirlikli_deger += agirlik

            ws3.cell(row=row, column=kriter_sayisi + 2, value=round(toplam_agirlikli_deger, 2))

        wb.save(file_name)

    def createTable4(self, file_name):
        grades_file, _ = QFileDialog.getOpenFileName(self, 'Not Dosyasını Seç', '',
                                                     'Excel Dosyaları (*.xlsx);;Tüm Dosyalar (*)')
        if not grades_file:
            QMessageBox.warning(self, "Hata", "Not dosyasını seçmediniz. Lütfen bir not dosyası seçin!")
            return

        try:
            self._generate_table4(grades_file, file_name)
            QMessageBox.information(self, "Başarı", "Tablo 4 başarıyla oluşturuldu ve kaydedildi.")
        except Exception as e:
            QMessageBox.critical(self, "Hata", f"Tablo 4 oluşturulurken bir hata oluştu: {e}")

    def _generate_table4(self, grades_file, file_name):
        grades_wb = load_workbook(grades_file, data_only=True)# Not dosyasını yükle
        weights_wb = load_workbook(file_name)

        grades_ws = grades_wb.active
        weights_ws = weights_wb['Tablo 3']  # Tablo 3 sayfasını seç

        if 'Tablo 4' in weights_wb.sheetnames:
            del weights_wb['Tablo 4']
        table4_ws = weights_wb.create_sheet('Tablo 4')

        # Ders çıktılarını ve ağırlıklarını saklamak için sözlük
        weights = {}
        max_values = {}
        for row in range(2, weights_ws.max_row + 1):
            ders_cikti = weights_ws.cell(row=row, column=1).value
            row_weights = {
                'Odev1': weights_ws.cell(row=row, column=2).value or 0,
                'Odev2': weights_ws.cell(row=row, column=3).value or 0,
                'Quiz': weights_ws.cell(row=row, column=4).value or 0,
                'Vize': weights_ws.cell(row=row, column=5).value or 0,
                'Final': weights_ws.cell(row=row, column=6).value or 0
            }
            row_weights = {key: float(value) if isinstance(value, (int, float)) else 0 for key, value in
                           row_weights.items()}
            weights[ders_cikti] = row_weights
            max_values[ders_cikti] = (weights_ws.cell(row=row, column=7).value or 0) * 100

        headers = ['Ders Çıktı', 'Odev1', 'Odev2', 'Quiz', 'Vize', 'Final', 'TOPLAM', 'MAX', '% Başarı']

        current_row = 1
        for student_row in range(2, grades_ws.max_row + 1):
            student_no = grades_ws.cell(row=student_row, column=1).value # Öğrenci numarasını al

            # Öğrenci başlığını ekle
            table4_ws.cell(row=current_row, column=1, value="TABLO 4")
            table4_ws.cell(row=current_row, column=2, value=f"Öğrenci {student_no} için")
            current_row += 1

            for col, header in enumerate(headers, 1):
                table4_ws.cell(row=current_row, column=col, value=header)
            current_row += 1

            grades = {
                'Odev1': grades_ws.cell(row=student_row, column=2).value or 0,
                'Odev2': grades_ws.cell(row=student_row, column=3).value or 0,
                'Quiz': grades_ws.cell(row=student_row, column=4).value or 0,
                'Vize': grades_ws.cell(row=student_row, column=5).value or 0,
                'Final': grades_ws.cell(row=student_row, column=6).value or 0
            }
            grades = {key: float(value) if isinstance(value, (int, float)) else 0 for key, value in grades.items()}

            # Ders çıktıları için hesaplamaları yap ve tabloya yaz
            for ders_cikti in weights.keys():
                weighted_total = 0
                for category, grade in grades.items():
                    weighted_total += grade * weights[ders_cikti][category]

                # Hesaplanan değerleri hücrelere yaz
                table4_ws.cell(row=current_row, column=1, value=ders_cikti)
                table4_ws.cell(row=current_row, column=2,
                               value=round(grades['Odev1'] * weights[ders_cikti]['Odev1'], 2))
                table4_ws.cell(row=current_row, column=3,
                               value=round(grades['Odev2'] * weights[ders_cikti]['Odev2'], 2))
                table4_ws.cell(row=current_row, column=4, value=round(grades['Quiz'] * weights[ders_cikti]['Quiz'], 2))
                table4_ws.cell(row=current_row, column=5, value=round(grades['Vize'] * weights[ders_cikti]['Vize'], 2))
                table4_ws.cell(row=current_row, column=6,
                               value=round(grades['Final'] * weights[ders_cikti]['Final'], 2))
                table4_ws.cell(row=current_row, column=7, value=round(weighted_total, 2))
                table4_ws.cell(row=current_row, column=8, value=round(max_values[ders_cikti], 2))

                # Başarı yüzdesini hesapla ve yaz
                success_rate = (weighted_total / max_values[ders_cikti] * 100) if max_values[ders_cikti] > 0 else 0
                table4_ws.cell(row=current_row, column=9, value=round(success_rate, 1))

                current_row += 1

            current_row += 2

        weights_wb.save(file_name)

    def createTable5(self, file_name):
        try:
            self._generate_table5(file_name)
            QMessageBox.information(self, "Başarı", "Tablo 5 başarıyla oluşturuldu ve kaydedildi.")
        except Exception as e:
            QMessageBox.critical(self, "Hata", f"Tablo 5 oluşturulurken bir hata oluştu: {e}")

    def _generate_table5(self, file_name):
        wb = load_workbook(file_name)

        table1_ws = wb["Tablo 1"]
        table4_ws = wb["Tablo 4"]

        if "Tablo 5" in wb.sheetnames:
            del wb["Tablo 5"]
        table5_ws = wb.create_sheet("Tablo 5")

        current_row = 1
        row = 1

        # Tablo 4'teki her öğrenci için tablo oluştur
        while row <= table4_ws.max_row:
            student_header = table4_ws.cell(row=row, column=2).value
            if student_header and "Öğrenci" in student_header:
                table5_ws.cell(row=current_row, column=1, value="TABLO 5")
                table5_ws.cell(row=current_row, column=2, value=student_header)
                current_row += 1

                # "Ders çıktısı" ve "Başarı Oranı" sütun başlıklarını ekle
                table5_ws.cell(row=current_row, column=2, value="Ders çıktısı")
                table5_ws.cell(row=current_row, column=8, value="Başarı Oranı")
                current_row += 1

                # Başarı oranlarını için sözlük oluştur
                success_rates = {}
                row += 2

                while row <= table4_ws.max_row and not (table4_ws.cell(row=row, column=2).value and "Öğrenci" in str(
                        table4_ws.cell(row=row, column=2).value)):
                    dc = table4_ws.cell(row=row, column=1).value
                    if dc and dc.startswith("DC"):
                        success_rate = table4_ws.cell(row=row, column=9).value
                        success_rates[dc] = success_rate
                    row += 1

                for prog_row in range(2, table1_ws.max_row + 1):
                    prog_output = table1_ws.cell(row=prog_row, column=1).value
                    table5_ws.cell(row=current_row, column=1, value=prog_output)

                    dc_success_total = 0
                    relation_evaluation = 0

                    for dc_col in range(2, table1_ws.max_column):
                        relation = table1_ws.cell(row=prog_row, column=dc_col).value
                        if str(relation) == "1":
                            dc_number = f"DC{dc_col - 1}"
                            success = success_rates.get(dc_number, 0)
                            table5_ws.cell(row=current_row, column=dc_col, value=success)
                            dc_success_total += success
                            relation_evaluation += 1

                        else:
                            table5_ws.cell(row=current_row, column=dc_col, value=0)

                    if relation_evaluation > 0:
                        avg_success = (dc_success_total / relation_evaluation)
                    else:
                        avg_success = 0
                    table5_ws.cell(row=current_row, column=8, value=round(avg_success, 1))

                    current_row += 1
                current_row += 2
            else:
                row += 1

        wb.save(file_name)

if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = Proje3GUI()
    window.show()
    sys.exit(app.exec_())
