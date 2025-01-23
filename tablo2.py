from openpyxl import Workbook, load_workbook
from PyQt5.QtWidgets import (
    QApplication, QWidget, QVBoxLayout, QLabel, QPushButton, QMessageBox, QTableWidget, 
    QTableWidgetItem, QHeaderView, QInputDialog, QFormLayout, QLineEdit, QSpinBox
)
from PyQt5.QtCore import Qt
import sys

class Tablo2App(QWidget):
    def __init__(self):
        super().__init__()
        self.initUI()

    def initUI(self):
        # Pencere başlığını ve boyutunu ayarla
        self.setWindowTitle("Tablo 2 Oluşturma ve Hesaplama")
        self.setGeometry(100, 100, 800, 600)

        # Ana dikey yerleşim düzenini oluştur
        self.layout = QVBoxLayout()

        # Değerlendirme kriterleri form düzeni
        self.kriter_layout = QFormLayout()
        self.kriter_label = QLabel('Değerlendirme Kriterleri ve Ağırlıkları:')
        self.kriter_layout.addRow(self.kriter_label)

        # Değerlendirme kriterleri ve ağırlıkları için giriş alanları oluştur
        self.kriter_inputs = []
        self.agirlik_inputs = []
        for i in range(5):
            # Kriter başlık alanı (sadece okunabilir)
            kriter_input = QLineEdit(f'Kriter {i + 1}')
            kriter_input.setStyleSheet(
                "background-color: rgb(230, 230, 250); color: black; border-radius: 5px; padding: 5px;"
            )
            kriter_input.setReadOnly(True)

            # Ağırlık giriş kutusu (0-100 aralığında)
            agirlik_input = QSpinBox()
            agirlik_input.setStyleSheet(
                "background-color: rgb(255, 255, 255); color: black; border-radius: 5px; padding: 5px;"
            )
            agirlik_input.setRange(0, 100)
            agirlik_input.setValue(20)  # Varsayılan değer olarak %20 ayarla
            agirlik_input.valueChanged.connect(self.updatePercentages)

            self.kriter_layout.addRow(kriter_input, agirlik_input)
            self.kriter_inputs.append(kriter_input)
            self.agirlik_inputs.append(agirlik_input)

        self.layout.addLayout(self.kriter_layout)

        # Tabloyu oluştur (3 satır, 7 sütun)
        self.table = QTableWidget(3, 7)
        self.table.setHorizontalHeaderLabels(["Ders Çıktı", "Ödev1", "Ödev2", "Quiz", "Vize", "Final", "Toplam"])
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.table.itemChanged.connect(self.validate_and_update)
        self.layout.addWidget(self.table)

        # İlk sütundaki hücreleri düzenlenemez hale getir
        for row in range(self.table.rowCount()):
            self.table.setItem(row, 0, QTableWidgetItem(f"DC{row + 1}"))
            self.table.item(row, 0).setFlags(Qt.ItemIsEnabled)

        # Satır ekleme butonu
        self.add_row_button = QPushButton("Satır Ekle")
        self.add_row_button.clicked.connect(self.addRow)
        self.layout.addWidget(self.add_row_button)

        # Tabloyu kaydetme butonu
        self.save_button = QPushButton("Tabloyu Kaydet")
        self.save_button.clicked.connect(self.saveTable)
        self.layout.addWidget(self.save_button)

        self.setLayout(self.layout)

        # Kriter ağırlıklarını tutan sözlük (varsayılan %20'lik ağırlıklar)
        self.percentages = {"Ödev1": 20, "Ödev2": 20, "Quiz": 20, "Vize": 20, "Final": 20}

    def addRow(self):
        # Yeni bir satır eklemek için fonksiyon
        current_row_count = self.table.rowCount()
        self.table.insertRow(current_row_count)
        self.table.setItem(current_row_count, 0, QTableWidgetItem(f"DC{current_row_count + 1}"))
        self.table.item(current_row_count, 0).setFlags(Qt.ItemIsEnabled)

    def updatePercentages(self):
        toplam_yuzde = sum(agirlik_input.value() for agirlik_input in self.agirlik_inputs)
        if toplam_yuzde > 100:
            QMessageBox.warning(self, "Hata", "Yüzdelerin toplamı 100 olmalıdır!")
            return

        for col, agirlik_input in enumerate(self.agirlik_inputs, start=1):
            self.percentages[f"Kriter {col}"] = agirlik_input.value()
            self.table.setHorizontalHeaderItem(col, QTableWidgetItem(f"Kriter {col} (%{agirlik_input.value()})"))
    def updatePercentages(self):
        toplam_yuzde = sum(agirlik_input.value() for agirlik_input in self.agirlik_inputs)

        if toplam_yuzde > 100:
            QMessageBox.warning(self, "Hata", "Yüzdelerin toplamı 100 olmalıdır!")
            return

        for col, (key, agirlik_input) in enumerate(zip(self.percentages.keys(), self.agirlik_inputs), start=1):
            self.percentages[key] = agirlik_input.value()
            self.table.setHorizontalHeaderItem(col, QTableWidgetItem(f"{key} (%{agirlik_input.value()})"))

    def validate_and_update(self, item):
        # Kullanıcının girdisini kontrol et
        if item.column() == 0 or item.column() == self.table.columnCount() - 1:
            return

        try:
            # Girilen değer sadece 0 veya 1 olmalı
            value = int(item.text())
            if value not in [0, 1]:
                raise ValueError
        except ValueError:
            QMessageBox.warning(self, "Hata", "Lütfen sadece 0 veya 1 girin.")
            item.setText("")  # Hatalı giriş yapıldığında temizleniyor
            return

        self.updateTotal(item.row())

    def updateTotal(self, row):
        # Toplam sütununu güncelleme fonksiyonu
        total = 0
        for col in range(1, self.table.columnCount() - 1):
            cell_item = self.table.item(row, col)
            if cell_item and cell_item.text():
                total += int(cell_item.text())

        # Toplam değerini ilgili sütuna yaz
        self.table.setItem(row, self.table.columnCount() - 1, QTableWidgetItem(str(total)))

    def saveTable(self):
        # Tabloyu Excel dosyasına kaydetme fonksiyonu
        file_name = "Tablo1.xlsx"

        try:
            try:
                # Var olan Excel dosyasını aç, yoksa yeni bir dosya oluştur
                wb = load_workbook(file_name)
            except FileNotFoundError:
                wb = Workbook()

            # Eğer "Tablo 2" sayfası varsa kaldır ve yeniden oluştur
            if "Tablo 2" in wb.sheetnames:
                ws2 = wb["Tablo 2"]
                wb.remove(ws2)
            ws2 = wb.create_sheet("Tablo 2")

            # Başlık satırlarını yaz
            row_count = self.table.rowCount()
            col_count = self.table.columnCount()

            for col in range(col_count):
                header_item = self.table.horizontalHeaderItem(col)
                ws2.cell(row=1, column=col + 1, value=header_item.text() if header_item else "")

            # Tablodaki verileri Excel sayfasına yaz
            for row in range(row_count):
                for col in range(col_count):
                    item = self.table.item(row, col)
                    value = item.text() if item else ""
                    ws2.cell(row=row + 2, column=col + 1, value=value)

            # Excel dosyasını kaydet
            wb.save(file_name)
            QMessageBox.information(self, "Başarı", f"Tablo 2 {file_name} dosyasına kaydedildi.")
        except Exception as e:
            QMessageBox.critical(self, "Hata", f"Tablo kaydedilirken bir hata oluştu: {e}")

